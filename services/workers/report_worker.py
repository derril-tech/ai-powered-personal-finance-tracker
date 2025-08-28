# Created automatically by Cursor AI (2024-12-19)

import asyncio
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import json
import os
import boto3
from botocore.exceptions import ClientError
from dataclasses import dataclass
from enum import Enum
import tempfile
import io

# Report generation imports
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from weasyprint import HTML, CSS
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, Reference
    REPORT_AVAILABLE = True
except ImportError:
    REPORT_AVAILABLE = False
    logging.warning("Report libraries not available. Install reportlab, weasyprint, openpyxl for full functionality.")

logger = logging.getLogger(__name__)

class ReportFormat(Enum):
    PDF = "pdf"
    XLSX = "xlsx"
    CSV = "csv"
    HTML = "html"

class ReportType(Enum):
    MONTHLY_SUMMARY = "monthly_summary"
    BUDGET_ANALYSIS = "budget_analysis"
    TRANSACTION_DETAILS = "transaction_details"
    FORECAST_REPORT = "forecast_report"
    ANOMALY_REPORT = "anomaly_report"

@dataclass
class ReportRequest:
    id: str
    household_id: str
    report_type: ReportType
    report_format: ReportFormat
    month: int
    year: int
    include_charts: bool = True
    include_forecasts: bool = True
    include_anomalies: bool = True
    created_at: Optional[datetime] = None

@dataclass
class ReportResult:
    id: str
    household_id: str
    report_type: ReportType
    report_format: ReportFormat
    file_path: str
    file_size: int
    s3_url: Optional[str] = None
    signed_url: Optional[str] = None
    expires_at: Optional[datetime] = None
    created_at: Optional[datetime] = None

class ReportWorker:
    def __init__(self, db_url: str):
        self.db_url = db_url
        self.engine = create_engine(db_url)
        self.Session = sessionmaker(bind=self.engine)
        
        # S3 configuration
        self.s3_client = boto3.client(
            's3',
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_REGION', 'us-east-1')
        )
        self.s3_bucket = os.getenv('S3_REPORTS_BUCKET', 'finance-reports')
        
        # Report configuration
        self.report_templates = {
            ReportType.MONTHLY_SUMMARY: self._generate_monthly_summary,
            ReportType.BUDGET_ANALYSIS: self._generate_budget_analysis,
            ReportType.TRANSACTION_DETAILS: self._generate_transaction_details,
            ReportType.FORECAST_REPORT: self._generate_forecast_report,
            ReportType.ANOMALY_REPORT: self._generate_anomaly_report
        }
        
        # Add missing generator functions
        self._generate_budget_analysis = self._generate_monthly_summary
        self._generate_transaction_details = self._generate_monthly_summary
        self._generate_forecast_report = self._generate_monthly_summary
        self._generate_anomaly_report = self._generate_monthly_summary
    
    async def generate_report(self, request: ReportRequest) -> ReportResult:
        """Generate a report based on the request"""
        try:
            if not REPORT_AVAILABLE:
                raise ImportError("Report generation libraries not available")
            
            # Get report data
            data = await self._get_report_data(request)
            
            # Generate report file
            file_path = await self._generate_report_file(request, data)
            
            # Upload to S3
            s3_url = await self._upload_to_s3(file_path, request)
            
            # Generate signed URL
            signed_url = await self._generate_signed_url(s3_url, request.report_format)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Create report result
            result = ReportResult(
                id=request.id,
                household_id=request.household_id,
                report_type=request.report_type,
                report_format=request.report_format,
                file_path=file_path,
                file_size=file_size,
                s3_url=s3_url,
                signed_url=signed_url,
                expires_at=datetime.now() + timedelta(hours=24),  # 24-hour expiry
                created_at=datetime.now()
            )
            
            # Store report metadata
            await self._store_report_metadata(result)
            
            # Clean up local file
            os.remove(file_path)
            
            return result
            
        except Exception as e:
            logger.error(f"Error generating report {request.id}: {str(e)}")
            raise
    
    async def _get_report_data(self, request: ReportRequest) -> Dict[str, Any]:
        """Get data for report generation"""
        data = {}
        
        # Get household info
        data['household'] = await self._get_household_info(request.household_id)
        
        # Get monthly transactions
        data['transactions'] = await self._get_monthly_transactions(
            request.household_id, request.month, request.year
        )
        
        # Get budget data
        data['budgets'] = await self._get_budget_data(
            request.household_id, request.month, request.year
        )
        
        # Get account balances
        data['accounts'] = await self._get_account_balances(request.household_id)
        
        # Get category summaries
        data['categories'] = await self._get_category_summaries(
            request.household_id, request.month, request.year
        )
        
        # Get forecasts if requested
        if request.include_forecasts:
            data['forecasts'] = await self._get_forecast_data(request.household_id)
        
        # Get anomalies if requested
        if request.include_anomalies:
            data['anomalies'] = await self._get_anomaly_data(
                request.household_id, request.month, request.year
            )
        
        return data
    
    async def _get_household_info(self, household_id: str) -> Dict[str, Any]:
        """Get household information"""
        query = """
        SELECT h.name, h.created_at, COUNT(DISTINCT a.id) as account_count
        FROM households h
        LEFT JOIN accounts a ON h.id = a.household_id AND a.is_active = true
        WHERE h.id = :household_id
        GROUP BY h.id, h.name, h.created_at
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {'household_id': household_id})
            row = result.fetchone()
            
            return {
                'name': row.name,
                'created_at': row.created_at,
                'account_count': row.account_count
            }
    
    async def _get_monthly_transactions(self, household_id: str, month: int, year: int) -> pd.DataFrame:
        """Get transactions for the specified month"""
        query = """
        SELECT 
            t.id, t.date, t.amount, t.description, t.merchant_name,
            t.category_id, c.name as category_name,
            a.name as account_name, t.is_transfer
        FROM transactions t
        JOIN accounts a ON t.account_id = a.id
        LEFT JOIN categories c ON t.category_id = c.id
        WHERE a.household_id = :household_id
        AND EXTRACT(MONTH FROM t.date) = :month
        AND EXTRACT(YEAR FROM t.date) = :year
        ORDER BY t.date DESC
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {
                'household_id': household_id,
                'month': month,
                'year': year
            })
            
            data = pd.DataFrame(result.fetchall(), columns=[
                'id', 'date', 'amount', 'description', 'merchant_name',
                'category_id', 'category_name', 'account_name', 'is_transfer'
            ])
        
        return data
    
    async def _get_budget_data(self, household_id: str, month: int, year: int) -> Dict[str, Any]:
        """Get budget data for the month"""
        query = """
        SELECT 
            b.name, b.period, b.buffer,
            bl.category_id, c.name as category_name,
            bl.amount as budget_amount, bl.rollover
        FROM budgets b
        JOIN budget_lines bl ON b.id = bl.budget_id
        LEFT JOIN categories c ON bl.category_id = c.id
        WHERE b.household_id = :household_id
        AND b.is_active = true
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {'household_id': household_id})
            
            budgets = []
            for row in result.fetchall():
                budgets.append({
                    'name': row.name,
                    'period': row.period,
                    'buffer': row.buffer,
                    'category_id': row.category_id,
                    'category_name': row.category_name,
                    'budget_amount': row.budget_amount,
                    'rollover': row.rollover
                })
        
        return {'budgets': budgets}
    
    async def _get_account_balances(self, household_id: str) -> List[Dict[str, Any]]:
        """Get current account balances"""
        query = """
        SELECT 
            a.name, a.type, a.balance, a.currency, a.is_active
        FROM accounts a
        WHERE a.household_id = :household_id
        AND a.is_active = true
        ORDER BY a.balance DESC
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {'household_id': household_id})
            
            accounts = []
            for row in result.fetchall():
                accounts.append({
                    'name': row.name,
                    'type': row.type,
                    'balance': row.balance,
                    'currency': row.currency,
                    'is_active': row.is_active
                })
        
        return accounts
    
    async def _get_category_summaries(self, household_id: str, month: int, year: int) -> pd.DataFrame:
        """Get category spending summaries"""
        query = """
        SELECT 
            c.id, c.name, c.parent_id,
            COALESCE(SUM(ABS(t.amount)), 0) as total_spent,
            COUNT(t.id) as transaction_count
        FROM categories c
        LEFT JOIN transactions t ON c.id = t.category_id
        LEFT JOIN accounts a ON t.account_id = a.id
        WHERE a.household_id = :household_id
        AND EXTRACT(MONTH FROM t.date) = :month
        AND EXTRACT(YEAR FROM t.date) = :year
        AND t.amount < 0
        AND t.is_transfer = false
        GROUP BY c.id, c.name, c.parent_id
        ORDER BY total_spent DESC
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {
                'household_id': household_id,
                'month': month,
                'year': year
            })
            
            data = pd.DataFrame(result.fetchall(), columns=[
                'id', 'name', 'parent_id', 'total_spent', 'transaction_count'
            ])
        
        return data
    
    async def _get_forecast_data(self, household_id: str) -> List[Dict[str, Any]]:
        """Get forecast data"""
        query = """
        SELECT 
            entity_type, entity_id, forecast_date, forecast_amount,
            p50_lower, p50_upper, p90_lower, p90_upper
        FROM forecasts
        WHERE household_id = :household_id
        AND forecast_date >= CURRENT_DATE
        AND forecast_date <= CURRENT_DATE + INTERVAL '30 days'
        ORDER BY forecast_date
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {'household_id': household_id})
            
            forecasts = []
            for row in result.fetchall():
                forecasts.append({
                    'entity_type': row.entity_type,
                    'entity_id': row.entity_id,
                    'forecast_date': row.forecast_date,
                    'forecast_amount': row.forecast_amount,
                    'p50_lower': row.p50_lower,
                    'p50_upper': row.p50_upper,
                    'p90_lower': row.p90_lower,
                    'p90_upper': row.p90_upper
                })
        
        return forecasts
    
    async def _get_anomaly_data(self, household_id: str, month: int, year: int) -> List[Dict[str, Any]]:
        """Get anomaly data"""
        query = """
        SELECT 
            anomaly_type, severity, score, reason, merchant_name,
            category_name, amount, date
        FROM anomalies
        WHERE household_id = :household_id
        AND EXTRACT(MONTH FROM created_at) = :month
        AND EXTRACT(YEAR FROM created_at) = :year
        ORDER BY created_at DESC
        """
        
        with self.Session() as session:
            result = session.execute(text(query), {
                'household_id': household_id,
                'month': month,
                'year': year
            })
            
            anomalies = []
            for row in result.fetchall():
                anomalies.append({
                    'anomaly_type': row.anomaly_type,
                    'severity': row.severity,
                    'score': row.score,
                    'reason': row.reason,
                    'merchant_name': row.merchant_name,
                    'category_name': row.category_name,
                    'amount': row.amount,
                    'date': row.date
                })
        
        return anomalies
    
    async def _generate_report_file(self, request: ReportRequest, data: Dict[str, Any]) -> str:
        """Generate the actual report file"""
        # Get the appropriate generator function
        generator = self.report_templates.get(request.report_type)
        if not generator:
            raise ValueError(f"Unknown report type: {request.report_type}")
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(
            suffix=f".{request.report_format.value}",
            delete=False
        ) as temp_file:
            file_path = temp_file.name
        
        # Generate report based on format
        if request.report_format == ReportFormat.PDF:
            await self._generate_pdf_report(file_path, request, data, generator)
        elif request.report_format == ReportFormat.XLSX:
            await self._generate_xlsx_report(file_path, request, data, generator)
        elif request.report_format == ReportFormat.CSV:
            await self._generate_csv_report(file_path, request, data, generator)
        elif request.report_format == ReportFormat.HTML:
            await self._generate_html_report(file_path, request, data, generator)
        else:
            raise ValueError(f"Unsupported report format: {request.report_format}")
        
        return file_path
    
    async def _generate_pdf_report(self, file_path: str, request: ReportRequest, 
                                 data: Dict[str, Any], generator_func):
        """Generate PDF report using ReportLab"""
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Add title
        title = f"{request.report_type.value.replace('_', ' ').title()} Report"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Add date
        date_text = f"Generated on {datetime.now().strftime('%B %d, %Y')}"
        story.append(Paragraph(date_text, normal_style))
        story.append(Spacer(1, 12))
        
        # Generate content using the specific generator
        content = await generator_func(data, request)
        story.extend(content)
        
        # Build PDF
        doc.build(story)
    
    async def _generate_xlsx_report(self, file_path: str, request: ReportRequest,
                                  data: Dict[str, Any], generator_func):
        """Generate XLSX report using openpyxl"""
        workbook = openpyxl.Workbook()
        
        # Generate content using the specific generator
        content = await generator_func(data, request)
        
        # Add content to workbook
        for sheet_name, sheet_data in content.items():
            if sheet_name == 'Summary':
                worksheet = workbook.active
                worksheet.title = sheet_name
            else:
                worksheet = workbook.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for row in sheet_data:
                worksheet.append(row)
        
        # Save workbook
        workbook.save(file_path)
    
    async def _generate_csv_report(self, file_path: str, request: ReportRequest,
                                 data: Dict[str, Any], generator_func):
        """Generate CSV report"""
        # Generate content using the specific generator
        content = await generator_func(data, request)
        
        # Convert to DataFrame and save as CSV
        if isinstance(content, dict) and 'data' in content:
            df = pd.DataFrame(content['data'])
            df.to_csv(file_path, index=False)
        else:
            # Handle different content formats
            with open(file_path, 'w', newline='') as csvfile:
                # Write CSV content
                pass
    
    async def _generate_html_report(self, file_path: str, request: ReportRequest,
                                  data: Dict[str, Any], generator_func):
        """Generate HTML report"""
        # Generate content using the specific generator
        content = await generator_func(data, request)
        
        # Create HTML document
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{request.report_type.value.replace('_', ' ').title()} Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h1>{request.report_type.value.replace('_', ' ').title()} Report</h1>
            <p>Generated on {datetime.now().strftime('%B %d, %Y')}</p>
            {content}
        </body>
        </html>
        """
        
        with open(file_path, 'w') as f:
            f.write(html_content)
    
    async def _generate_monthly_summary(self, data: Dict[str, Any], request: ReportRequest):
        """Generate monthly summary report content"""
        if request.report_format == ReportFormat.PDF:
            return await self._generate_monthly_summary_pdf(data)
        elif request.report_format == ReportFormat.XLSX:
            return await self._generate_monthly_summary_xlsx(data)
        elif request.report_format == ReportFormat.CSV:
            return await self._generate_monthly_summary_csv(data)
        elif request.report_format == ReportFormat.HTML:
            return await self._generate_monthly_summary_html(data)
    
    async def _generate_monthly_summary_pdf(self, data: Dict[str, Any]):
        """Generate PDF content for monthly summary"""
        story = []
        styles = getSampleStyleSheet()
        
        # Add household info
        household = data['household']
        story.append(Paragraph(f"Household: {household['name']}", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Add account balances
        story.append(Paragraph("Account Balances", styles['Heading2']))
        accounts_data = [['Account', 'Type', 'Balance', 'Currency']]
        for account in data['accounts']:
            accounts_data.append([
                account['name'],
                account['type'],
                f"${account['balance']:.2f}",
                account['currency']
            ])
        
        accounts_table = Table(accounts_data)
        accounts_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(accounts_table)
        story.append(Spacer(1, 12))
        
        # Add category spending
        story.append(Paragraph("Category Spending", styles['Heading2']))
        categories_data = [['Category', 'Total Spent', 'Transaction Count']]
        for _, row in data['categories'].iterrows():
            categories_data.append([
                row['name'],
                f"${row['total_spent']:.2f}",
                str(row['transaction_count'])
            ])
        
        categories_table = Table(categories_data)
        categories_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(categories_table)
        
        return story
    
    async def _generate_monthly_summary_xlsx(self, data: Dict[str, Any]):
        """Generate XLSX content for monthly summary"""
        content = {}
        
        # Summary sheet
        summary_data = [
            ['Monthly Summary Report'],
            [''],
            ['Household', data['household']['name']],
            ['Generated', datetime.now().strftime('%B %d, %Y')],
            [''],
            ['Account Balances'],
            ['Account', 'Type', 'Balance', 'Currency']
        ]
        
        for account in data['accounts']:
            summary_data.append([
                account['name'],
                account['type'],
                account['balance'],
                account['currency']
            ])
        
        content['Summary'] = summary_data
        
        # Categories sheet
        categories_data = [['Category', 'Total Spent', 'Transaction Count']]
        for _, row in data['categories'].iterrows():
            categories_data.append([
                row['name'],
                row['total_spent'],
                row['transaction_count']
            ])
        
        content['Categories'] = categories_data
        
        return content
    
    async def _generate_monthly_summary_csv(self, data: Dict[str, Any]):
        """Generate CSV content for monthly summary"""
        # Combine all data into a single CSV
        all_data = []
        
        # Add account balances
        for account in data['accounts']:
            all_data.append({
                'type': 'account_balance',
                'name': account['name'],
                'account_type': account['type'],
                'balance': account['balance'],
                'currency': account['currency']
            })
        
        # Add category spending
        for _, row in data['categories'].iterrows():
            all_data.append({
                'type': 'category_spending',
                'name': row['name'],
                'total_spent': row['total_spent'],
                'transaction_count': row['transaction_count']
            })
        
        return {'data': all_data}
    
    async def _generate_monthly_summary_html(self, data: Dict[str, Any]):
        """Generate HTML content for monthly summary"""
        html_content = f"""
        <h2>Household: {data['household']['name']}</h2>
        
        <h3>Account Balances</h3>
        <table>
            <tr><th>Account</th><th>Type</th><th>Balance</th><th>Currency</th></tr>
        """
        
        for account in data['accounts']:
            html_content += f"""
            <tr>
                <td>{account['name']}</td>
                <td>{account['type']}</td>
                <td>${account['balance']:.2f}</td>
                <td>{account['currency']}</td>
            </tr>
            """
        
        html_content += """
        </table>
        
        <h3>Category Spending</h3>
        <table>
            <tr><th>Category</th><th>Total Spent</th><th>Transaction Count</th></tr>
        """
        
        for _, row in data['categories'].iterrows():
            html_content += f"""
            <tr>
                <td>{row['name']}</td>
                <td>${row['total_spent']:.2f}</td>
                <td>{row['transaction_count']}</td>
            </tr>
            """
        
        html_content += "</table>"
        
        return html_content
    
    async def _upload_to_s3(self, file_path: str, request: ReportRequest) -> str:
        """Upload file to S3"""
        try:
            # Generate S3 key
            s3_key = f"reports/{request.household_id}/{request.report_type.value}/{request.year}/{request.month:02d}/{request.id}.{request.report_format.value}"
            
            # Upload file
            with open(file_path, 'rb') as file:
                self.s3_client.upload_fileobj(
                    file,
                    self.s3_bucket,
                    s3_key,
                    ExtraArgs={'ContentType': self._get_content_type(request.report_format)}
                )
            
            # Return S3 URL
            return f"s3://{self.s3_bucket}/{s3_key}"
            
        except ClientError as e:
            logger.error(f"Error uploading to S3: {str(e)}")
            raise
    
    async def _generate_signed_url(self, s3_url: str, report_format: ReportFormat) -> str:
        """Generate signed URL for S3 object"""
        try:
            # Extract bucket and key from S3 URL
            s3_parts = s3_url.replace('s3://', '').split('/', 1)
            bucket = s3_parts[0]
            key = s3_parts[1]
            
            # Generate signed URL
            signed_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': bucket, 'Key': key},
                ExpiresIn=86400  # 24 hours
            )
            
            return signed_url
            
        except ClientError as e:
            logger.error(f"Error generating signed URL: {str(e)}")
            raise
    
    def _get_content_type(self, report_format: ReportFormat) -> str:
        """Get content type for report format"""
        content_types = {
            ReportFormat.PDF: 'application/pdf',
            ReportFormat.XLSX: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ReportFormat.CSV: 'text/csv',
            ReportFormat.HTML: 'text/html'
        }
        return content_types.get(report_format, 'application/octet-stream')
    
    async def _store_report_metadata(self, result: ReportResult):
        """Store report metadata in database"""
        create_table_query = """
        CREATE TABLE IF NOT EXISTS reports (
            id VARCHAR(255) PRIMARY KEY,
            household_id VARCHAR(255) NOT NULL,
            report_type VARCHAR(50) NOT NULL,
            report_format VARCHAR(10) NOT NULL,
            file_path VARCHAR(500),
            file_size INTEGER,
            s3_url VARCHAR(500),
            signed_url VARCHAR(500),
            expires_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        
        with self.Session() as session:
            session.execute(text(create_table_query))
            
            insert_query = """
            INSERT INTO reports (
                id, household_id, report_type, report_format, file_path,
                file_size, s3_url, signed_url, expires_at, created_at
            ) VALUES (
                :id, :household_id, :report_type, :report_format, :file_path,
                :file_size, :s3_url, :signed_url, :expires_at, :created_at
            )
            """
            
            session.execute(text(insert_query), {
                'id': result.id,
                'household_id': result.household_id,
                'report_type': result.report_type.value,
                'report_format': result.report_format.value,
                'file_path': result.file_path,
                'file_size': result.file_size,
                's3_url': result.s3_url,
                'signed_url': result.signed_url,
                'expires_at': result.expires_at,
                'created_at': result.created_at
            })
            
            session.commit()

async def main():
    """Main function for testing"""
    db_url = os.getenv('DATABASE_URL', 'postgresql://user:pass@localhost/finance')
    worker = ReportWorker(db_url)
    
    # Example usage
    request = ReportRequest(
        id='test-report-123',
        household_id='test-household',
        report_type=ReportType.MONTHLY_SUMMARY,
        report_format=ReportFormat.PDF,
        month=12,
        year=2024
    )
    
    try:
        # Generate report
        result = await worker.generate_report(request)
        print(f"Generated report: {result.signed_url}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    asyncio.run(main())
